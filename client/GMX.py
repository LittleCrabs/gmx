# GMX Client - 主程序
import sys
import os
import re
import json
import email
import imaplib
import requests
import pandas as pd
from datetime import datetime, timedelta
from email.header import decode_header
from email.utils import parsedate_to_datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# PySide6 GUI
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QTabWidget, QMessageBox, QProgressBar, QFileDialog, QStackedWidget,
    QGroupBox, QFormLayout, QCheckBox, QSpinBox, QHeaderView, QDialog,
    QDialogButtonBox, QComboBox
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont, QIcon

# ==================== 配置 ====================
DEFAULT_SERVER_URL = "http://localhost:5050"
CONFIG_FILE = "gmx_config.json"
EXCEL_FILE = "billing_records.xlsx"
PRINTER_REPORT_FILE = "Printer_Status_Report.xlsx"


def load_local_config():
    """加载本地配置"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except:
            pass
    return {'server_url': DEFAULT_SERVER_URL}


def save_local_config(config):
    """保存本地配置"""
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f)

# ==================== 样式 ====================
TITLE_FILL = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
URGENT_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BLACK_FONT = Font(color="000000", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14, name="Courier New")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Courier New")


# ==================== API客户端 ====================
class APIClient:
    def __init__(self):
        self.token = None
        self.user = None
        self.server_url = DEFAULT_SERVER_URL
    
    def set_server(self, url):
        self.server_url = url.rstrip('/')
    
    def _headers(self):
        if self.token:
            return {'Authorization': f'Bearer {self.token}', 'Content-Type': 'application/json'}
        return {'Content-Type': 'application/json'}
    
    def login(self, username, password):
        try:
            resp = requests.post(f'{self.server_url}/api/login', 
                               json={'username': username, 'password': password}, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                self.token = data['token']
                self.user = data['user']
                return True, data['user']
            return False, resp.json().get('error', 'Login failed')
        except Exception as e:
            return False, str(e)
    
    def register(self, username, password):
        try:
            resp = requests.post(f'{self.server_url}/api/register',
                               json={'username': username, 'password': password}, timeout=10)
            if resp.status_code == 201:
                return True, 'Registration successful'
            return False, resp.json().get('error', 'Registration failed')
        except Exception as e:
            return False, str(e)
    
    def get_client_config(self):
        try:
            resp = requests.get(f'{self.server_url}/api/client/config', headers=self._headers(), timeout=10)
            if resp.status_code == 200:
                return True, resp.json()
            return False, resp.json().get('error', 'Failed to get config')
        except Exception as e:
            return False, str(e)
    
    # 管理员API
    def get_users(self):
        resp = requests.get(f'{self.server_url}/api/admin/users', headers=self._headers(), timeout=10)
        return resp.json() if resp.status_code == 200 else []
    
    def create_user(self, data):
        resp = requests.post(f'{self.server_url}/api/admin/users', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 201
    
    def update_user(self, user_id, data):
        resp = requests.put(f'{self.server_url}/api/admin/users/{user_id}', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 200
    
    def delete_user(self, user_id):
        resp = requests.delete(f'{self.server_url}/api/admin/users/{user_id}', headers=self._headers(), timeout=10)
        return resp.status_code == 200
    
    def get_emails(self):
        resp = requests.get(f'{self.server_url}/api/admin/emails', headers=self._headers(), timeout=10)
        return resp.json() if resp.status_code == 200 else []
    
    def add_email(self, data):
        resp = requests.post(f'{self.server_url}/api/admin/emails', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 201
    
    def update_email(self, email_id, data):
        resp = requests.put(f'{self.server_url}/api/admin/emails/{email_id}', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 200
    
    def delete_email(self, email_id):
        resp = requests.delete(f'{self.server_url}/api/admin/emails/{email_id}', headers=self._headers(), timeout=10)
        return resp.status_code == 200
    
    def get_imap(self):
        resp = requests.get(f'{self.server_url}/api/admin/imap', headers=self._headers(), timeout=10)
        return resp.json() if resp.status_code == 200 else None
    
    def save_imap(self, data):
        resp = requests.post(f'{self.server_url}/api/admin/imap', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 200
    
    def save_config(self, data):
        resp = requests.post(f'{self.server_url}/api/admin/config', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 200
    
    # SN映射API
    def get_sn_mappings(self):
        resp = requests.get(f'{self.server_url}/api/sn-mappings', headers=self._headers(), timeout=10)
        return resp.json() if resp.status_code == 200 else []
    
    def add_sn_mapping(self, data):
        resp = requests.post(f'{self.server_url}/api/sn-mappings', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 201
    
    def update_sn_mapping(self, mapping_id, data):
        resp = requests.put(f'{self.server_url}/api/sn-mappings/{mapping_id}', headers=self._headers(), json=data, timeout=10)
        return resp.status_code == 200
    
    def delete_sn_mapping(self, mapping_id):
        resp = requests.delete(f'{self.server_url}/api/sn-mappings/{mapping_id}', headers=self._headers(), timeout=10)
        return resp.status_code == 200


api = APIClient()


# ==================== 邮件处理 ====================
def decode_mime_header(header_value):
    if header_value is None:
        return ""
    decoded_parts = decode_header(header_value)
    result = []
    for part, charset in decoded_parts:
        if isinstance(part, bytes):
            result.append(part.decode(charset or 'utf-8', errors='ignore'))
        else:
            result.append(part)
    return ''.join(result)


def get_email_body(msg):
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))
            if content_type == "text/plain" and "attachment" not in content_disposition:
                try:
                    charset = part.get_content_charset() or 'utf-8'
                    payload = part.get_payload(decode=True)
                    if payload:
                        body = payload.decode(charset, errors='ignore')
                        break
                except:
                    continue
    else:
        try:
            charset = msg.get_content_charset() or 'utf-8'
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode(charset, errors='ignore')
        except:
            pass
    return body


def parse_body(text):
    def extract_single(pattern):
        m = re.search(pattern, text, re.MULTILINE)
        return m.group(1).strip() if m else None

    def extract_block(start_label):
        pattern = rf"\[{re.escape(start_label)}\]\s*\n(.*?)(?=\n\[|$)"
        m = re.search(pattern, text, re.DOTALL)
        return m.group(1).strip() if m else None

    def extract_consumables(text):
        consumables_list = []
        consumables_text = extract_block("Consumables")
        if consumables_text:
            if "There are no errors to be reported" in consumables_text:
                return []
            pattern = r"(.+?)\n\s*<([^>]+)>"
            matches = re.findall(pattern, consumables_text)
            if not matches:
                pattern = r"(.+?)<([^>]+)>"
                matches = re.findall(pattern, consumables_text.replace('\n', ' '))
            for match in matches:
                if len(match) == 2:
                    consumables_list.append((match[0].strip(), match[1].strip()))
        return consumables_list

    return {
        "product_name": extract_single(r"Product Name\s+(.+)"),
        "ipv4_address": extract_single(r"IPv4 Address\s+([\d.]+)"),
        "serial_number": extract_single(r"Serial Number\s+(\d+)"),
        "billing_meter_1": int(extract_single(r"Billing Meter\[1\]\s+(\d+)") or 0),
        "billing_meter_2": int(extract_single(r"Billing Meter\[2\]\s+(\d+)") or 0),
        "billing_meter_3": int(extract_single(r"Billing Meter\[3\]\s+(\d+)") or 0),
        "billing_meter_4": int(extract_single(r"Billing Meter\[4\]\s+(\d+)") or 0),
        "billing_meter_5": int(extract_single(r"Billing Meter\[5\]\s+(\d+)") or 0),
        "consumables": extract_block("Consumables"),
        "consumables_trouble": extract_block("Consumables Trouble"),
        "consumables_list": extract_consumables(text)
    }


# ==================== 邮件获取线程 ====================
class FetchEmailThread(QThread):
    progress = Signal(str)
    progress_count = Signal(int, int)  # current, total
    finished_signal = Signal(bool, str, object)
    
    def __init__(self, config):
        super().__init__()
        self.config = config
    
    def run(self):
        try:
            imap_cfg = self.config['imap']
            sender_emails = self.config['sender_emails']
            interval = self.config['interval']
            
            self.progress.emit("Connecting to IMAP server...")
            mail = imaplib.IMAP4_SSL(imap_cfg['server'], imap_cfg['port'])
            mail.login(imap_cfg['username'], imap_cfg['password'])
            mail.select('INBOX')
            
            since_date = (datetime.now() - timedelta(days=interval)).strftime("%d-%b-%Y")
            self.progress.emit(f"Searching emails since {since_date}...")
            
            # 先收集所有邮件ID
            all_email_ids = []
            for sender in sender_emails:
                search_criteria = f'(FROM "{sender}" SINCE "{since_date}")'
                status, data = mail.search(None, search_criteria)
                if status == 'OK' and data[0]:
                    all_email_ids.extend(data[0].split())
            
            total = len(all_email_ids)
            if total == 0:
                mail.close()
                mail.logout()
                self.finished_signal.emit(True, "No emails found", None)
                return
            
            self.progress.emit(f"Found {total} emails, fetching...")
            self.progress_count.emit(0, total)
            
            records = []
            for idx, email_id in enumerate(all_email_ids):
                self.progress_count.emit(idx + 1, total)
                self.progress.emit(f"Processing {idx + 1}/{total}...")
                
                status, msg_data = mail.fetch(email_id, '(RFC822)')
                if status != 'OK':
                    continue
                raw_email = msg_data[0][1]
                msg = email.message_from_bytes(raw_email)
                body = get_email_body(msg)
                parsed = parse_body(body)
                parsed["message_id"] = email_id.decode()
                parsed["mail_date"] = msg.get('Date', '')
                records.append(parsed)
            
            mail.close()
            mail.logout()
            
            if records:
                df = pd.DataFrame(records)
                df.to_excel(EXCEL_FILE, index=False)
                self.create_printer_report(df)
                self.finished_signal.emit(True, f"Completed: {len(records)} records", df)
            else:
                self.finished_signal.emit(True, "No valid emails found", None)
        except Exception as e:
            self.finished_signal.emit(False, str(e), None)
    
    def create_printer_report(self, df):
        def parse_mail_date(date_str):
            try:
                return parsedate_to_datetime(date_str)
            except:
                return datetime.min
        
        df['parsed_date'] = df['mail_date'].apply(parse_mail_date)
        df_sorted = df.sort_values('parsed_date', ascending=False)
        latest_records = []
        seen_devices = set()
        for _, row in df_sorted.iterrows():
            key = (row['product_name'], row['serial_number'])
            if key not in seen_devices:
                seen_devices.add(key)
                latest_records.append(row)
        
        if not latest_records:
            return
        
        wb = Workbook()
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        for i, record in enumerate(latest_records):
            sn = str(record['serial_number'])[:31]
            ws = wb.active if i == 0 else wb.create_sheet(title=sn)
            if i == 0:
                ws.title = sn
            
            row = 1
            # Device Info
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row=row, column=1, value="Device Information").font = TITLE_FONT
            ws.cell(row=row, column=1).fill = TITLE_FILL
            ws.cell(row=row, column=1).alignment = center_align
            row += 1
            
            for label, value in [("Product Name", record['product_name']),
                                ("IPv4 Address", record['ipv4_address']),
                                ("Serial Number", record['serial_number'])]:
                ws.cell(row=row, column=1, value=label).fill = DATA_FILL
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
                ws.cell(row=row, column=2, value=value)
                row += 1
            row += 1
            
            # Consumables
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row=row, column=1, value="Consumables Status").font = TITLE_FONT
            ws.cell(row=row, column=1).fill = TITLE_FILL
            row += 1
            
            for col, h in enumerate(["Item", "Status", "Priority"], 1):
                ws.cell(row=row, column=col, value=h).font = HEADER_FONT
                ws.cell(row=row, column=col).fill = HEADER_FILL
            row += 1
            
            consumables = record.get('consumables_list', [])
            if consumables:
                for item, status in consumables:
                    priority = "Urgent" if "Replace Now" in status else ("Warning" if "Replace Soon" in status else "Normal")
                    ws.cell(row=row, column=1, value=item).fill = DATA_FILL
                    status_cell = ws.cell(row=row, column=2, value=status)
                    if "Replace Now" in status:
                        status_cell.fill = URGENT_FILL
                        status_cell.font = WHITE_FONT
                    elif "Replace Soon" in status:
                        status_cell.fill = WARN_FILL
                    else:
                        status_cell.fill = DATA_FILL
                    ws.cell(row=row, column=3, value=priority).fill = DATA_FILL
                    row += 1
            else:
                ws.cell(row=row, column=1, value="No errors").fill = DATA_FILL
                row += 1
            row += 1
            
            # Billing Meters
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.cell(row=row, column=1, value="Billing Meters").font = TITLE_FONT
            ws.cell(row=row, column=1).fill = TITLE_FILL
            row += 1
            
            for col, h in enumerate(["Meter", "Count", "Description"], 1):
                ws.cell(row=row, column=col, value=h).font = HEADER_FONT
                ws.cell(row=row, column=col).fill = HEADER_FILL
            row += 1
            
            for meter, count, desc in [
                ("Billing Meter [1]", record['billing_meter_1'], "B&W Prints"),
                ("Billing Meter [2]", record['billing_meter_2'], "Color Prints"),
                ("Billing Meter [3]", record['billing_meter_3'], "Large Format"),
                ("Billing Meter [4]", record['billing_meter_4'], "Total Count"),
                ("Billing Meter [5]", record['billing_meter_5'], "Other")
            ]:
                ws.cell(row=row, column=1, value=meter).fill = DATA_FILL
                ws.cell(row=row, column=2, value=f"{count:,}").fill = DATA_FILL
                ws.cell(row=row, column=3, value=desc).fill = DATA_FILL
                row += 1
        
        for ws in wb.worksheets:
            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 15
        
        wb.save(PRINTER_REPORT_FILE)


# ==================== 登录窗口 ====================
class LoginWindow(QWidget):
    login_success = Signal(dict)
    
    def __init__(self):
        super().__init__()
        # 加载保存的服务器地址
        config = load_local_config()
        api.set_server(config.get('server_url', DEFAULT_SERVER_URL))
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("GMX - Login")
        self.setFixedSize(350, 250)
        
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(40, 30, 40, 30)
        
        # Title
        title = QLabel("GMX Billing Meter")
        title.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        # Form
        form = QFormLayout()
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Username")
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Password")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        form.addRow("Username:", self.username_input)
        form.addRow("Password:", self.password_input)
        layout.addLayout(form)
        
        # Buttons
        btn_layout = QHBoxLayout()
        self.login_btn = QPushButton("Login")
        self.login_btn.clicked.connect(self.do_login)
        self.register_btn = QPushButton("Register")
        self.register_btn.clicked.connect(self.do_register)
        btn_layout.addWidget(self.login_btn)
        btn_layout.addWidget(self.register_btn)
        layout.addLayout(btn_layout)
        
        # Status
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)
        
        self.setLayout(layout)
        self.password_input.returnPressed.connect(self.do_login)
    
    def do_login(self):
        username = self.username_input.text().strip()
        password = self.password_input.text()
        if not username or not password:
            self.status_label.setText("Please enter username and password")
            return
        
        self.status_label.setText("Logging in...")
        success, result = api.login(username, password)
        if success:
            self.login_success.emit(result)
        else:
            self.status_label.setText(f"Error: {result}")
    
    def do_register(self):
        username = self.username_input.text().strip()
        password = self.password_input.text()
        if not username or not password:
            self.status_label.setText("Please enter username and password")
            return
        
        success, msg = api.register(username, password)
        self.status_label.setText(msg if success else f"Error: {msg}")


# ==================== 主窗口 ====================
class MainWindow(QMainWindow):
    def __init__(self, user):
        super().__init__()
        self.user = user
        self.config = None
        self.init_ui()
        self.load_config()
    
    def init_ui(self):
        self.setWindowTitle(f"GMX Billing Meter - {self.user['username']}")
        self.setMinimumSize(900, 600)
        
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        
        # Tabs
        self.tabs = QTabWidget()
        
        # Tab 1: Fetch Emails
        self.tabs.addTab(self.create_fetch_tab(), "Fetch Emails")
        
        # Tab 2: SN Mapping (所有用户可用)
        self.tabs.addTab(self.create_sn_mapping_tab(), "SN Mapping")
        
        # Admin tabs
        if self.user.get('is_admin'):
            self.tabs.addTab(self.create_users_tab(), "Users")
            self.tabs.addTab(self.create_emails_tab(), "Sender Emails")
            self.tabs.addTab(self.create_imap_tab(), "IMAP Config")
            self.tabs.addTab(self.create_settings_tab(), "Settings")
        
        layout.addWidget(self.tabs)
    
    def create_fetch_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Info
        info_group = QGroupBox("Configuration")
        info_layout = QFormLayout()
        self.imap_label = QLabel("Loading...")
        self.emails_label = QLabel("Loading...")
        self.interval_label = QLabel("Loading...")
        info_layout.addRow("IMAP Server:", self.imap_label)
        info_layout.addRow("Sender Emails:", self.emails_label)
        info_layout.addRow("Interval (days):", self.interval_label)
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # Buttons
        btn_layout = QHBoxLayout()
        self.fetch_btn = QPushButton("Fetch Emails")
        self.fetch_btn.clicked.connect(self.fetch_emails)
        self.refresh_btn = QPushButton("Refresh Config")
        self.refresh_btn.clicked.connect(self.load_config)
        self.open_excel_btn = QPushButton("Open Report")
        self.open_excel_btn.clicked.connect(self.open_report)
        btn_layout.addWidget(self.fetch_btn)
        btn_layout.addWidget(self.refresh_btn)
        btn_layout.addWidget(self.open_excel_btn)
        layout.addLayout(btn_layout)
        
        # Progress
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        layout.addWidget(self.progress_bar)
        
        self.progress_label = QLabel("")
        layout.addWidget(self.progress_label)
        
        # Search
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search Serial:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Enter serial number...")
        self.search_input.textChanged.connect(self.filter_results)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)
        
        # Results table
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(8)
        self.results_table.setHorizontalHeaderLabels(["Product", "Serial", "IP", "B&W", "Color", "Meter3", "Meter4", "Remark"])
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.results_table)
        
        # 保存原始数据用于搜索
        self.results_data = None
        
        return widget
    
    def create_sn_mapping_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Add Mapping")
        add_btn.clicked.connect(self.add_sn_mapping)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_sn_mappings)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.sn_table = QTableWidget()
        self.sn_table.setColumnCount(4)
        self.sn_table.setHorizontalHeaderLabels(["ID", "Serial Number", "Company Name", "Actions"])
        self.sn_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.sn_table)
        
        return widget
    
    def create_users_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Add User")
        add_btn.clicked.connect(self.add_user)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_users)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(5)
        self.users_table.setHorizontalHeaderLabels(["ID", "Username", "Admin", "Active", "Actions"])
        self.users_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.users_table)
        
        return widget
    
    def create_emails_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Add Email")
        add_btn.clicked.connect(self.add_email)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_emails)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.emails_table = QTableWidget()
        self.emails_table.setColumnCount(4)
        self.emails_table.setHorizontalHeaderLabels(["ID", "Email", "Active", "Actions"])
        self.emails_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.emails_table)
        
        return widget
    
    def create_imap_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        form = QFormLayout()
        self.imap_server = QLineEdit()
        self.imap_port = QSpinBox()
        self.imap_port.setRange(1, 65535)
        self.imap_port.setValue(993)
        self.imap_username = QLineEdit()
        self.imap_password = QLineEdit()
        self.imap_password.setEchoMode(QLineEdit.EchoMode.Password)
        
        form.addRow("Server:", self.imap_server)
        form.addRow("Port:", self.imap_port)
        form.addRow("Username:", self.imap_username)
        form.addRow("Password:", self.imap_password)
        layout.addLayout(form)
        
        save_btn = QPushButton("Save IMAP Config")
        save_btn.clicked.connect(self.save_imap)
        layout.addWidget(save_btn)
        layout.addStretch()
        
        return widget
    
    def create_settings_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 服务器设置
        server_group = QGroupBox("Server Settings")
        server_layout = QFormLayout()
        self.server_url_input = QLineEdit()
        config = load_local_config()
        self.server_url_input.setText(config.get('server_url', DEFAULT_SERVER_URL))
        server_layout.addRow("Server URL:", self.server_url_input)
        server_group.setLayout(server_layout)
        layout.addWidget(server_group)
        
        save_server_btn = QPushButton("Save Server (requires re-login)")
        save_server_btn.clicked.connect(self.save_server_url)
        layout.addWidget(save_server_btn)
        
        # 系统设置
        system_group = QGroupBox("System Settings")
        system_layout = QFormLayout()
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(1, 365)
        self.interval_spin.setValue(7)
        system_layout.addRow("Fetch Interval (days):", self.interval_spin)
        system_group.setLayout(system_layout)
        layout.addWidget(system_group)
        
        save_btn = QPushButton("Save Settings")
        save_btn.clicked.connect(self.save_settings)
        layout.addWidget(save_btn)
        layout.addStretch()
        
        return widget
    
    def save_server_url(self):
        new_url = self.server_url_input.text().strip()
        if not new_url:
            QMessageBox.warning(self, "Error", "Server URL cannot be empty")
            return
        config = load_local_config()
        config['server_url'] = new_url
        save_local_config(config)
        QMessageBox.information(self, "Success", "Server URL saved. Please restart and re-login to apply.")

    
    def load_config(self):
        success, result = api.get_client_config()
        if success:
            self.config = result
            imap = result.get('imap')
            if imap:
                self.imap_label.setText(f"{imap['server']}:{imap['port']}")
            emails = result.get('sender_emails', [])
            self.emails_label.setText(f"{len(emails)} configured")
            self.interval_label.setText(str(result.get('interval', 7)))
            
            # 加载SN映射
            self.load_sn_mappings()
            
            if self.user.get('is_admin'):
                self.load_users()
                self.load_emails()
                self.load_imap()
                self.interval_spin.setValue(result.get('interval', 7))
        else:
            QMessageBox.warning(self, "Error", f"Failed to load config: {result}")
    
    def load_sn_mappings(self):
        mappings = api.get_sn_mappings()
        self.sn_table.setRowCount(len(mappings))
        for i, m in enumerate(mappings):
            self.sn_table.setItem(i, 0, QTableWidgetItem(str(m['id'])))
            self.sn_table.setItem(i, 1, QTableWidgetItem(m['serial_number']))
            self.sn_table.setItem(i, 2, QTableWidgetItem(m['company_name']))
            
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)
            del_btn = QPushButton("Delete")
            del_btn.clicked.connect(lambda _, mid=m['id']: self.delete_sn_mapping(mid))
            btn_layout.addWidget(del_btn)
            self.sn_table.setCellWidget(i, 3, btn_widget)
    
    def add_sn_mapping(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add SN Mapping")
        layout = QFormLayout(dialog)
        
        sn_input = QLineEdit()
        company_input = QLineEdit()
        
        layout.addRow("Serial Number:", sn_input)
        layout.addRow("Company Name:", company_input)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            sn = sn_input.text().strip()
            company = company_input.text().strip()
            if sn and company:
                if api.add_sn_mapping({'serial_number': sn, 'company_name': company}):
                    self.load_sn_mappings()
                    self.load_config()  # 刷新config中的sn_mappings
                else:
                    QMessageBox.warning(self, "Error", "Failed to add mapping (SN may already exist)")
            else:
                QMessageBox.warning(self, "Error", "Please fill in all fields")
    
    def delete_sn_mapping(self, mapping_id):
        if QMessageBox.question(self, "Confirm", "Delete this mapping?") == QMessageBox.StandardButton.Yes:
            if api.delete_sn_mapping(mapping_id):
                self.load_sn_mappings()
                self.load_config()
    
    def load_users(self):
        users = api.get_users()
        self.users_table.setRowCount(len(users))
        for i, u in enumerate(users):
            self.users_table.setItem(i, 0, QTableWidgetItem(str(u['id'])))
            self.users_table.setItem(i, 1, QTableWidgetItem(u['username']))
            self.users_table.setItem(i, 2, QTableWidgetItem("Yes" if u['is_admin'] else "No"))
            self.users_table.setItem(i, 3, QTableWidgetItem("Yes" if u['is_active'] else "No"))
            
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)
            del_btn = QPushButton("Delete")
            del_btn.clicked.connect(lambda _, uid=u['id']: self.delete_user(uid))
            btn_layout.addWidget(del_btn)
            self.users_table.setCellWidget(i, 4, btn_widget)
    
    def load_emails(self):
        emails = api.get_emails()
        self.emails_table.setRowCount(len(emails))
        for i, e in enumerate(emails):
            self.emails_table.setItem(i, 0, QTableWidgetItem(str(e['id'])))
            self.emails_table.setItem(i, 1, QTableWidgetItem(e['email']))
            self.emails_table.setItem(i, 2, QTableWidgetItem("Yes" if e['is_active'] else "No"))
            
            btn_widget = QWidget()
            btn_layout = QHBoxLayout(btn_widget)
            btn_layout.setContentsMargins(2, 2, 2, 2)
            del_btn = QPushButton("Delete")
            del_btn.clicked.connect(lambda _, eid=e['id']: self.delete_email(eid))
            btn_layout.addWidget(del_btn)
            self.emails_table.setCellWidget(i, 3, btn_widget)
    
    def load_imap(self):
        imap = api.get_imap()
        if imap:
            self.imap_server.setText(imap.get('server', ''))
            self.imap_port.setValue(imap.get('port', 993))
            self.imap_username.setText(imap.get('username', ''))
            self.imap_password.setText(imap.get('password', ''))
    
    def add_user(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add User")
        layout = QFormLayout(dialog)
        
        username = QLineEdit()
        password = QLineEdit()
        password.setEchoMode(QLineEdit.EchoMode.Password)
        is_admin = QCheckBox()
        
        layout.addRow("Username:", username)
        layout.addRow("Password:", password)
        layout.addRow("Admin:", is_admin)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            if api.create_user({'username': username.text(), 'password': password.text(), 'is_admin': is_admin.isChecked()}):
                self.load_users()
            else:
                QMessageBox.warning(self, "Error", "Failed to create user")
    
    def delete_user(self, user_id):
        if QMessageBox.question(self, "Confirm", "Delete this user?") == QMessageBox.StandardButton.Yes:
            if api.delete_user(user_id):
                self.load_users()
    
    def add_email(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Email")
        layout = QFormLayout(dialog)
        
        email_input = QLineEdit()
        desc = QLineEdit()
        
        layout.addRow("Email:", email_input)
        layout.addRow("Description:", desc)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            if api.add_email({'email': email_input.text(), 'description': desc.text()}):
                self.load_emails()
            else:
                QMessageBox.warning(self, "Error", "Failed to add email")
    
    def delete_email(self, email_id):
        if QMessageBox.question(self, "Confirm", "Delete this email?") == QMessageBox.StandardButton.Yes:
            if api.delete_email(email_id):
                self.load_emails()
    
    def save_imap(self):
        data = {
            'server': self.imap_server.text(),
            'port': self.imap_port.value(),
            'username': self.imap_username.text(),
            'password': self.imap_password.text()
        }
        if api.save_imap(data):
            QMessageBox.information(self, "Success", "IMAP config saved")
            self.load_config()
        else:
            QMessageBox.warning(self, "Error", "Failed to save IMAP config")
    
    def save_settings(self):
        if api.save_config({'interval': str(self.interval_spin.value())}):
            QMessageBox.information(self, "Success", "Settings saved")
            self.load_config()
        else:
            QMessageBox.warning(self, "Error", "Failed to save settings")
    
    def fetch_emails(self):
        if not self.config or not self.config.get('imap'):
            QMessageBox.warning(self, "Error", "No IMAP configuration")
            return
        
        self.fetch_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.progress_label.setText("Starting...")
        
        self.fetch_thread = FetchEmailThread(self.config)
        self.fetch_thread.progress.connect(lambda msg: self.progress_label.setText(msg))
        self.fetch_thread.progress_count.connect(self.update_progress)
        self.fetch_thread.finished_signal.connect(self.on_fetch_complete)
        self.fetch_thread.start()
    
    def update_progress(self, current, total):
        if total > 0:
            self.progress_bar.setMaximum(total)
            self.progress_bar.setValue(current)
    
    def on_fetch_complete(self, success, message, df):
        self.fetch_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.progress_label.setText(message)
        
        if success and df is not None:
            self.results_data = df  # 保存数据用于搜索
            self.display_results(df)
            QMessageBox.information(self, "Success", f"Saved to {PRINTER_REPORT_FILE}")
        elif not success:
            QMessageBox.warning(self, "Error", message)
    
    def display_results(self, df):
        # 获取SN映射
        sn_map = self.config.get('sn_mappings', {}) if self.config else {}
        
        self.results_table.setRowCount(len(df))
        for idx, (_, row) in enumerate(df.iterrows()):
            sn = str(row.get('serial_number', ''))
            remark = sn_map.get(sn, '')  # 根据SN查找公司名
            
            self.results_table.setItem(idx, 0, QTableWidgetItem(str(row.get('product_name', ''))))
            self.results_table.setItem(idx, 1, QTableWidgetItem(sn))
            self.results_table.setItem(idx, 2, QTableWidgetItem(str(row.get('ipv4_address', ''))))
            self.results_table.setItem(idx, 3, QTableWidgetItem(str(row.get('billing_meter_1', 0))))
            self.results_table.setItem(idx, 4, QTableWidgetItem(str(row.get('billing_meter_2', 0))))
            self.results_table.setItem(idx, 5, QTableWidgetItem(str(row.get('billing_meter_3', 0))))
            self.results_table.setItem(idx, 6, QTableWidgetItem(str(row.get('billing_meter_4', 0))))
            self.results_table.setItem(idx, 7, QTableWidgetItem(remark))
    
    def filter_results(self, text):
        if self.results_data is None:
            return
        
        if not text.strip():
            self.display_results(self.results_data)
        else:
            # 过滤包含搜索文本的行
            filtered = self.results_data[
                self.results_data['serial_number'].astype(str).str.contains(text, case=False, na=False)
            ]
            self.display_results(filtered)
    
    def open_report(self):
        if os.path.exists(PRINTER_REPORT_FILE):
            os.startfile(PRINTER_REPORT_FILE)
        else:
            QMessageBox.warning(self, "Error", "Report file not found")


# ==================== 应用入口 ====================
class GMXApp(QStackedWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("GMX Billing Meter")
        
        self.login_window = LoginWindow()
        self.login_window.login_success.connect(self.on_login_success)
        self.addWidget(self.login_window)
        
        self.main_window = None
    
    def on_login_success(self, user):
        self.main_window = MainWindow(user)
        self.addWidget(self.main_window)
        self.setCurrentWidget(self.main_window)
        self.setMinimumSize(900, 600)
        self.resize(1000, 700)


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    window = GMXApp()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

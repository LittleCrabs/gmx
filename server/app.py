from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import jwt
import datetime
import os
import re
import json
import email
import imaplib
import io
from email.header import decode_header
from email.utils import parsedate_to_datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
CORS(app)

# 配置
app.config['SECRET_KEY'] = 'gmx-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///gmx.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


# ==================== 数据模型 ====================
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class EmailConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    description = db.Column(db.String(256))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)


class IMAPConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    server = db.Column(db.String(120), nullable=False)
    port = db.Column(db.Integer, default=993)
    username = db.Column(db.String(120), nullable=False)
    password = db.Column(db.String(256), nullable=False)
    is_active = db.Column(db.Boolean, default=True)


class SystemConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(80), unique=True, nullable=False)
    value = db.Column(db.String(256), nullable=False)


class SNMapping(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    serial_number = db.Column(db.String(80), unique=True, nullable=False)
    company_name = db.Column(db.String(256), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)


class EmailRecord(db.Model):
    """缓存的邮件记录"""
    id = db.Column(db.Integer, primary_key=True)
    message_id = db.Column(db.String(100), unique=True, nullable=False)
    mail_date = db.Column(db.String(100))
    product_name = db.Column(db.String(200))
    serial_number = db.Column(db.String(80), index=True)  # 添加索引便于查询
    ipv4_address = db.Column(db.String(50))
    billing_meter_1 = db.Column(db.Integer, default=0)
    billing_meter_2 = db.Column(db.Integer, default=0)
    billing_meter_3 = db.Column(db.Integer, default=0)
    billing_meter_4 = db.Column(db.Integer, default=0)
    billing_meter_5 = db.Column(db.Integer, default=0)
    consumables = db.Column(db.Text)  # JSON格式存储耗材列表
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)
    


# ==================== 认证装饰器 ====================
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token is missing'}), 401
        try:
            token = token.replace('Bearer ', '')
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = User.query.get(data['user_id'])
            if not current_user or not current_user.is_active:
                return jsonify({'error': 'Invalid or inactive user'}), 401
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        return f(current_user, *args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token is missing'}), 401
        try:
            token = token.replace('Bearer ', '')
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = User.query.get(data['user_id'])
            if not current_user or not current_user.is_admin:
                return jsonify({'error': 'Admin access required'}), 403
        except:
            return jsonify({'error': 'Invalid token'}), 401
        return f(current_user, *args, **kwargs)
    return decorated


# ==================== 用户认证API ====================
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    user = User.query.filter_by(username=username).first()
    if user and user.check_password(password) and user.is_active:
        token = jwt.encode({
            'user_id': user.id,
            'is_admin': user.is_admin,
            'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        return jsonify({
            'token': token,
            'user': {
                'id': user.id,
                'username': user.username,
                'is_admin': user.is_admin
            }
        })
    return jsonify({'error': 'Invalid credentials'}), 401


@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Username already exists'}), 400
    
    user = User(username=username)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return jsonify({'message': 'User registered successfully'}), 201


@app.route('/api/user/info', methods=['GET'])
@token_required
def get_user_info(current_user):
    return jsonify({
        'id': current_user.id,
        'username': current_user.username,
        'is_admin': current_user.is_admin
    })


# ==================== 用户管理API (管理员) ====================
@app.route('/api/admin/users', methods=['GET'])
@admin_required
def get_users(current_user):
    users = User.query.all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'is_admin': u.is_admin,
        'is_active': u.is_active,
        'created_at': u.created_at.isoformat()
    } for u in users])


@app.route('/api/admin/users', methods=['POST'])
@admin_required
def create_user(current_user):
    data = request.get_json()
    if User.query.filter_by(username=data['username']).first():
        return jsonify({'error': 'Username already exists'}), 400
    
    user = User(
        username=data['username'],
        is_admin=data.get('is_admin', False),
        is_active=data.get('is_active', True)
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()
    return jsonify({'message': 'User created', 'id': user.id}), 201


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@admin_required
def update_user(current_user, user_id):
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    
    if 'is_admin' in data:
        user.is_admin = data['is_admin']
    if 'is_active' in data:
        user.is_active = data['is_active']
    if 'password' in data and data['password']:
        user.set_password(data['password'])
    
    db.session.commit()
    return jsonify({'message': 'User updated'})


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(current_user, user_id):
    if user_id == current_user.id:
        return jsonify({'error': 'Cannot delete yourself'}), 400
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'User deleted'})


# ==================== 邮箱配置API ====================
@app.route('/api/emails', methods=['GET'])
@token_required
def get_emails(current_user):
    emails = EmailConfig.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': e.id,
        'email': e.email,
        'description': e.description
    } for e in emails])


@app.route('/api/admin/emails', methods=['GET'])
@admin_required
def get_all_emails(current_user):
    emails = EmailConfig.query.all()
    return jsonify([{
        'id': e.id,
        'email': e.email,
        'description': e.description,
        'is_active': e.is_active,
        'created_at': e.created_at.isoformat()
    } for e in emails])


@app.route('/api/admin/emails', methods=['POST'])
@admin_required
def add_email(current_user):
    data = request.get_json()
    if EmailConfig.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Email already exists'}), 400
    
    email_config = EmailConfig(
        email=data['email'],
        description=data.get('description', ''),
        is_active=data.get('is_active', True)
    )
    db.session.add(email_config)
    db.session.commit()
    return jsonify({'message': 'Email added', 'id': email_config.id}), 201


@app.route('/api/admin/emails/<int:email_id>', methods=['PUT'])
@admin_required
def update_email(current_user, email_id):
    email_config = EmailConfig.query.get_or_404(email_id)
    data = request.get_json()
    
    if 'email' in data:
        email_config.email = data['email']
    if 'description' in data:
        email_config.description = data['description']
    if 'is_active' in data:
        email_config.is_active = data['is_active']
    
    db.session.commit()
    return jsonify({'message': 'Email updated'})


@app.route('/api/admin/emails/<int:email_id>', methods=['DELETE'])
@admin_required
def delete_email(current_user, email_id):
    email_config = EmailConfig.query.get_or_404(email_id)
    db.session.delete(email_config)
    db.session.commit()
    return jsonify({'message': 'Email deleted'})

# ==================== IMAP配置API ====================
@app.route('/api/imap', methods=['GET'])
@token_required
def get_imap_config(current_user):
    imap = IMAPConfig.query.filter_by(is_active=True).first()
    if imap:
        return jsonify({
            'server': imap.server,
            'port': imap.port,
            'username': imap.username,
            'password': imap.password
        })
    return jsonify({'error': 'No IMAP config found'}), 404


@app.route('/api/admin/imap', methods=['GET'])
@admin_required
def get_imap_config_admin(current_user):
    imap = IMAPConfig.query.first()
    if imap:
        return jsonify({
            'id': imap.id,
            'server': imap.server,
            'port': imap.port,
            'username': imap.username,
            'password': imap.password,
            'is_active': imap.is_active
        })
    return jsonify({'error': 'No IMAP config found'}), 404


@app.route('/api/admin/imap', methods=['POST', 'PUT'])
@admin_required
def save_imap_config(current_user):
    data = request.get_json()
    imap = IMAPConfig.query.first()
    
    if imap:
        imap.server = data['server']
        imap.port = data.get('port', 993)
        imap.username = data['username']
        imap.password = data['password']
        imap.is_active = data.get('is_active', True)
    else:
        imap = IMAPConfig(
            server=data['server'],
            port=data.get('port', 993),
            username=data['username'],
            password=data['password'],
            is_active=data.get('is_active', True)
        )
        db.session.add(imap)
    
    db.session.commit()
    return jsonify({'message': 'IMAP config saved'})


# ==================== 系统配置API ====================
@app.route('/api/config', methods=['GET'])
@token_required
def get_config(current_user):
    configs = SystemConfig.query.all()
    return jsonify({c.key: c.value for c in configs})


@app.route('/api/admin/config', methods=['POST'])
@admin_required
def save_config(current_user):
    data = request.get_json()
    for key, value in data.items():
        config = SystemConfig.query.filter_by(key=key).first()
        if config:
            config.value = str(value)
        else:
            config = SystemConfig(key=key, value=str(value))
            db.session.add(config)
    db.session.commit()
    return jsonify({'message': 'Config saved'})


# ==================== 客户端完整配置API ====================
@app.route('/api/client/config', methods=['GET'])
@token_required
def get_client_config(current_user):
    """获取客户端运行所需的全部配置"""
    # IMAP配置
    imap = IMAPConfig.query.filter_by(is_active=True).first()
    imap_config = None
    if imap:
        imap_config = {
            'server': imap.server,
            'port': imap.port,
            'username': imap.username,
            'password': imap.password
        }
    
    # 发件人邮箱列表
    emails = EmailConfig.query.filter_by(is_active=True).all()
    sender_emails = [e.email for e in emails]
    
    # 系统配置
    configs = SystemConfig.query.all()
    system_config = {c.key: c.value for c in configs}
    
    # SN映射
    sn_mappings = SNMapping.query.all()
    sn_map = {s.serial_number: s.company_name for s in sn_mappings}
    
    return jsonify({
        'imap': imap_config,
        'sender_emails': sender_emails,
        'interval': int(system_config.get('interval', 7)),
        'config': system_config,
        'sn_mappings': sn_map
    })


# ==================== SN映射API ====================
@app.route('/api/sn-mappings', methods=['GET'])
@token_required
def get_sn_mappings(current_user):
    mappings = SNMapping.query.all()
    return jsonify([{
        'id': m.id,
        'serial_number': m.serial_number,
        'company_name': m.company_name
    } for m in mappings])


@app.route('/api/sn-mappings', methods=['POST'])
@token_required
def add_sn_mapping(current_user):
    data = request.get_json()
    if SNMapping.query.filter_by(serial_number=data['serial_number']).first():
        return jsonify({'error': 'Serial number already exists'}), 400
    
    mapping = SNMapping(
        serial_number=data['serial_number'],
        company_name=data['company_name']
    )
    db.session.add(mapping)
    db.session.commit()
    return jsonify({'message': 'Mapping added', 'id': mapping.id}), 201


@app.route('/api/sn-mappings/<int:mapping_id>', methods=['PUT'])
@token_required
def update_sn_mapping(current_user, mapping_id):
    mapping = SNMapping.query.get_or_404(mapping_id)
    data = request.get_json()
    
    if 'serial_number' in data:
        mapping.serial_number = data['serial_number']
    if 'company_name' in data:
        mapping.company_name = data['company_name']
    
    db.session.commit()
    return jsonify({'message': 'Mapping updated'})


@app.route('/api/sn-mappings/<int:mapping_id>', methods=['DELETE'])
@token_required
def delete_sn_mapping(current_user, mapping_id):
    mapping = SNMapping.query.get_or_404(mapping_id)
    db.session.delete(mapping)
    db.session.commit()
    return jsonify({'message': 'Mapping deleted'})


# ==================== 邮件获取API ====================
def get_email_body(msg):
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition"))
            if content_type == "text/plain" and "attachment" not in content_disposition:
                try:
                    charset = part.get_content_charset() or 'utf-8'
                    payload = part.get_payload(decode=True)
                    if payload:
                        body = payload.decode(charset, errors='ignore')
                        break
                except:
                    continue
    else:
        try:
            charset = msg.get_content_charset() or 'utf-8'
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode(charset, errors='ignore')
        except:
            pass
    return body


def parse_email_body(text):
    def extract_single(pattern):
        m = re.search(pattern, text, re.MULTILINE)
        return m.group(1).strip() if m else None

    def extract_block(start_label):
        pattern = rf"\[{re.escape(start_label)}\]\s*\n(.*?)(?=\n\[|$)"
        m = re.search(pattern, text, re.DOTALL)
        return m.group(1).strip() if m else None

    def extract_consumables(text):
        consumables_list = []
        consumables_text = extract_block("Consumables")
        if consumables_text:
            if "There are no errors to be reported" in consumables_text:
                return []
            pattern = r"(.+?)\n\s*<([^>]+)>"
            matches = re.findall(pattern, consumables_text)
            if not matches:
                pattern = r"(.+?)<([^>]+)>"
                matches = re.findall(pattern, consumables_text.replace('\n', ' '))
            for match in matches:
                if len(match) == 2:
                    item = match[0].strip()
                    status = match[1].strip()
                    consumables_list.append([item, status])
        if not consumables_list:
            lines = text.split('\n')
            in_consumables_section = False
            for i, line in enumerate(lines):
                line = line.strip()
                if "Consumables" in line and "[" in line and "]" in line:
                    in_consumables_section = True
                    continue
                elif in_consumables_section and line.startswith("["):
                    break
                if in_consumables_section and line and not line.startswith("["):
                    status_match = re.search(r"<([^>]+)>", line)
                    if status_match:
                        status = status_match.group(1).strip()
                        item = line[:status_match.start()].strip()
                        if item:
                            consumables_list.append([item, status])
                    elif i+1 < len(lines) and re.search(r"<([^>]+)>", lines[i+1].strip()):
                        item = line
                        status_line = lines[i+1].strip()
                        status_match = re.search(r"<([^>]+)>", status_line)
                        if status_match:
                            status = status_match.group(1).strip()
                            consumables_list.append([item, status])
        return consumables_list

    return {
        "product_name": extract_single(r"Product Name\s+(.+)"),
        "ipv4_address": extract_single(r"IPv4 Address\s+([\d.]+)"),
        "serial_number": extract_single(r"Serial Number\s+(\d+)"),
        "billing_meter_1": int(extract_single(r"Billing Meter\[1\]\s+(\d+)") or 0),
        "billing_meter_2": int(extract_single(r"Billing Meter\[2\]\s+(\d+)") or 0),
        "billing_meter_3": int(extract_single(r"Billing Meter\[3\]\s+(\d+)") or 0),
        "billing_meter_4": int(extract_single(r"Billing Meter\[4\]\s+(\d+)") or 0),
        "billing_meter_5": int(extract_single(r"Billing Meter\[5\]\s+(\d+)") or 0),
        "consumables_list": extract_consumables(text),
    }


@app.route('/api/fetch-emails', methods=['POST'])
@token_required
def fetch_emails(current_user):
    """获取邮件，只获取新邮件并缓存"""
    try:
        data = request.get_json() or {}
        force_refresh = data.get('force', False)  # 是否强制刷新
        
        imap_cfg = IMAPConfig.query.filter_by(is_active=True).first()
        if not imap_cfg:
            return jsonify({'error': 'No IMAP configuration'}), 400
        
        sender_emails = [e.email for e in EmailConfig.query.filter_by(is_active=True).all()]
        if not sender_emails:
            return jsonify({'error': 'No sender emails configured'}), 400
        
        configs = {c.key: c.value for c in SystemConfig.query.all()}
        interval = int(configs.get('interval', 7))
        
        # 获取已缓存的message_id
        cached_ids = set(r.message_id for r in EmailRecord.query.all())
        print(f"[Fetch] Cached records: {len(cached_ids)}")
        
        # 连接IMAP
        print(f"[Fetch] Connecting to {imap_cfg.server}...")
        mail = imaplib.IMAP4_SSL(imap_cfg.server, imap_cfg.port)
        mail.login(imap_cfg.username, imap_cfg.password)
        mail.select('INBOX')
        
        since_date = (datetime.datetime.now() - timedelta(days=interval)).strftime("%d-%b-%Y")
        print(f"[Fetch] Searching emails since {since_date}...")
        
        # 收集所有邮件ID
        all_email_ids = []
        for sender in sender_emails:
            search_criteria = f'(FROM "{sender}" SINCE "{since_date}")'
            status, data = mail.search(None, search_criteria)
            if status == 'OK' and data[0]:
                all_email_ids.extend(data[0].split())
        
        # 过滤出新邮件
        new_email_ids = []
        for eid in all_email_ids:
            msg_id = eid.decode()
            if force_refresh or msg_id not in cached_ids:
                new_email_ids.append(eid)
        
        total = len(new_email_ids)
        print(f"[Fetch] Found {len(all_email_ids)} emails, {total} new to fetch")
        
        # 只获取新邮件
        new_count = 0
        for idx, email_id in enumerate(new_email_ids):
            if (idx + 1) % 10 == 0 or idx == 0:
                print(f"[Fetch] Processing {idx + 1}/{total}...")
            
            status, msg_data = mail.fetch(email_id, '(RFC822)')
            if status != 'OK':
                continue
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            body = get_email_body(msg)
            parsed = parse_email_body(body)
            
            msg_id = email_id.decode()
            
            # 保存到数据库
            record = EmailRecord(
                message_id=msg_id,
                mail_date=msg.get('Date', ''),
                product_name=parsed.get('product_name'),
                serial_number=parsed.get('serial_number'),
                ipv4_address=parsed.get('ipv4_address'),
                billing_meter_1=parsed.get('billing_meter_1', 0),
                billing_meter_2=parsed.get('billing_meter_2', 0),
                billing_meter_3=parsed.get('billing_meter_3', 0),
                billing_meter_4=parsed.get('billing_meter_4', 0),
                billing_meter_5=parsed.get('billing_meter_5', 0),
                consumables=json.dumps(parsed.get('consumables_list', []))
            )
            db.session.add(record)
            new_count += 1
        
        db.session.commit()
        mail.close()
        mail.logout()
        
        print(f"[Fetch] Added {new_count} new records")
        
        # 返回按serial_number去重的最新记录
        records = get_latest_records_by_sn()
        
        return jsonify({'records': records, 'count': len(records), 'new_count': new_count})
    except Exception as e:
        print(f"[Fetch] Error: {e}")
        return jsonify({'error': str(e)}), 500


def get_latest_records_by_sn():
    """获取每个serial_number的最新记录（返回字典列表）"""
    records = get_latest_email_records_by_sn()
    return [{
        'message_id': r.message_id,
        'mail_date': r.mail_date,
        'product_name': r.product_name,
        'serial_number': r.serial_number,
        'ipv4_address': r.ipv4_address,
        'billing_meter_1': r.billing_meter_1,
        'billing_meter_2': r.billing_meter_2,
        'billing_meter_3': r.billing_meter_3,
        'billing_meter_4': r.billing_meter_4,
        'billing_meter_5': r.billing_meter_5
    } for r in records]


def get_latest_email_records_by_sn():
    """获取每个serial_number的最新记录（返回EmailRecord对象列表）"""
    from sqlalchemy import func
    
    # 子查询：获取每个serial_number的最大id（最新记录）
    subquery = db.session.query(
        EmailRecord.serial_number,
        func.max(EmailRecord.id).label('max_id')
    ).filter(
        EmailRecord.serial_number.isnot(None)
    ).group_by(EmailRecord.serial_number).subquery()
    
    # 主查询：获取这些最新记录的完整信息
    records = db.session.query(EmailRecord).join(
        subquery,
        (EmailRecord.serial_number == subquery.c.serial_number) & 
        (EmailRecord.id == subquery.c.max_id)
    ).order_by(EmailRecord.mail_date.desc()).all()
    
    return records


@app.route('/api/email-records', methods=['GET'])
@token_required
def get_email_records(current_user):
    """获取按serial_number去重的最新邮件记录"""
    records = get_latest_records_by_sn()
    return jsonify(records)


# ==================== Excel导出API ====================
# 样式定义
TITLE_FILL = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DATA_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
URGENT_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BLACK_FONT = Font(color="000000", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14, name="Courier New")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Courier New")


@app.route('/api/export/billing-records', methods=['GET'])
@token_required
def export_billing_records(current_user):
    """导出按serial_number去重的最新数据Excel"""
    records = get_latest_email_records_by_sn()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing Records"
    
    # 表头
    headers = ['Product Name', 'Serial Number', 'IP Address', 'B&W', 'Color', 'Meter3', 'Meter4', 'Meter5', 'Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    # 数据
    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.product_name)
        ws.cell(row=row_idx, column=2, value=r.serial_number)
        ws.cell(row=row_idx, column=3, value=r.ipv4_address)
        ws.cell(row=row_idx, column=4, value=r.billing_meter_1)
        ws.cell(row=row_idx, column=5, value=r.billing_meter_2)
        ws.cell(row=row_idx, column=6, value=r.billing_meter_3)
        ws.cell(row=row_idx, column=7, value=r.billing_meter_4)
        ws.cell(row=row_idx, column=8, value=r.billing_meter_5)
        ws.cell(row=row_idx, column=9, value=r.mail_date)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['I'].width = 30
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'billing_records_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/api/export/printer-report', methods=['GET'])
@token_required
def export_printer_report(current_user):
    """导出打印机状态报告Excel（每个打印机一个sheet）"""
    latest_records = get_latest_email_records_by_sn()
    
    if not latest_records:
        return jsonify({'error': 'No data to export'}), 400
    
    wb = Workbook()
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    for i, record in enumerate(latest_records):
        sn = str(record.serial_number or 'Unknown')[:31]
        if i == 0:
            ws = wb.active
            ws.title = sn
        else:
            ws = wb.create_sheet(title=sn)
        
        current_row = 1
        
        # ========== Device Information ==========
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        title_cell = ws.cell(row=current_row, column=1, value="Device Information")
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = center_align
        current_row += 1
        
        device_info = [
            ("Product Name", record.product_name),
            ("IPv4 Address", record.ipv4_address),
            ("Serial Number", record.serial_number)
        ]
        for label, value in device_info:
            ws.cell(row=current_row, column=1, value=label).fill = DATA_FILL
            ws.cell(row=current_row, column=1).font = BLACK_FONT
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        
        current_row += 1
        
        # ========== Consumables Status ==========
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        title_cell = ws.cell(row=current_row, column=1, value="Consumables Status")
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = center_align
        current_row += 1
        
        headers = ["Item", "Status", "Priority"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = center_align
        current_row += 1
        
        # 解析consumables JSON
        consumables_data = []
        if record.consumables:
            try:
                consumables_data = json.loads(record.consumables)
            except:
                pass
        
        if consumables_data:
            for item_data in consumables_data:
                item = item_data[0] if len(item_data) > 0 else ''
                status = item_data[1] if len(item_data) > 1 else ''
                priority = "Urgent" if "Replace Now" in status else ("Warning" if "Replace Soon" in status else "Normal")
                
                item_cell = ws.cell(row=current_row, column=1, value=item)
                item_cell.fill = DATA_FILL
                
                status_cell = ws.cell(row=current_row, column=2, value=status)
                if "Replace Now" in status:
                    status_cell.fill = URGENT_FILL
                    status_cell.font = WHITE_FONT
                elif "Replace Soon" in status:
                    status_cell.fill = WARN_FILL
                    status_cell.font = BLACK_FONT
                else:
                    status_cell.fill = DATA_FILL
                
                priority_cell = ws.cell(row=current_row, column=3, value=priority)
                priority_cell.alignment = center_align
                if priority == "Urgent":
                    priority_cell.font = WHITE_FONT
                    priority_cell.fill = URGENT_FILL
                elif priority == "Warning":
                    priority_cell.font = BLACK_FONT
                    priority_cell.fill = WARN_FILL
                else:
                    priority_cell.fill = DATA_FILL
                
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="There are no errors to be reported.").fill = DATA_FILL
            ws.cell(row=current_row, column=2, value="Normal").fill = DATA_FILL
            ws.cell(row=current_row, column=3, value="Normal").fill = DATA_FILL
            current_row += 1
        
        current_row += 1
        
        # ========== Billing Meters ==========
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        title_cell = ws.cell(row=current_row, column=1, value="Billing Meters")
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = center_align
        current_row += 1
        
        headers = ["Meter", "Count", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = center_align
        current_row += 1
        
        billing_meters = [
            ("Billing Meter [1]", record.billing_meter_1, "B&W Prints"),
            ("Billing Meter [2]", record.billing_meter_2, "Color Prints"),
            ("Billing Meter [3]", record.billing_meter_3, "Large Format"),
            ("Billing Meter [4]", record.billing_meter_4, "Total Count"),
            ("Billing Meter [5]", record.billing_meter_5, "Other")
        ]
        for meter, count, desc in billing_meters:
            ws.cell(row=current_row, column=1, value=meter).fill = DATA_FILL
            count_cell = ws.cell(row=current_row, column=2, value=f"{count:,}" if count else "0")
            count_cell.fill = DATA_FILL
            count_cell.alignment = right_align
            ws.cell(row=current_row, column=3, value=desc).fill = DATA_FILL
            current_row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Printer_Status_Report_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# ==================== 初始化数据库 ====================
def init_db():
    with app.app_context():
        db.create_all()
        # 创建默认管理员
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', is_admin=True)
            admin.set_password('admin123')
            db.session.add(admin)
        
        # 创建默认IMAP配置
        if not IMAPConfig.query.first():
            imap = IMAPConfig(
                server='imap.dreamhost.com',
                port=993,
                username='autometer@archer.com.sg',
                password='bqr0ucp!ZMN5vjz.wme'
            )
            db.session.add(imap)
        
        # 创建默认发件人邮箱
        default_emails = ['scan@archer.com.sg']
        for email_addr in default_emails:
            if not EmailConfig.query.filter_by(email=email_addr).first():
                db.session.add(EmailConfig(email=email_addr, description='Default sender'))
        
        # 创建默认系统配置
        if not SystemConfig.query.filter_by(key='interval').first():
            db.session.add(SystemConfig(key='interval', value='7'))
        
        db.session.commit()


# 初始化数据库（gunicorn启动时也会执行）
init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5050, debug=True)
